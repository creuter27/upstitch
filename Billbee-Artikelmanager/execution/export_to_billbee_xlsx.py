"""
Export Google Sheet rows → Billbee-compatible XLSX for article import.

Why this script exists
----------------------
The Billbee REST API cannot update CustomFields, BillOfMaterial, TARIC Code,
Country of origin, or dimensions (LengthCm/WidthCm/HeightCm) via PATCH.
The only way to bulk-write these fields is Billbee's XLSX article import:
  Artikel → Importieren → Billbee XLSX

Column selection via ColumnsToUpload tab
-----------------------------------------
The script reads the 'ColumnsToUpload' tab (maintained by download_to_sheet.py)
and exports exactly the columns whose checkbox is TRUE.  Since the downloaded
tab now uses the same column names as the Billbee XLSX format, no mapping is
needed — column names pass through directly.

Internal-only columns (Type, Action, BOM_Count, StockCurrent, etc.) are never
exported even if checked.  BOM_SKUs is handled specially (see below).

Row selection via Google Sheets filter
---------------------------------------
Apply any filter to the 'upload' tab before running this script.
Only VISIBLE rows are included in the output XLSX.  Hidden rows are skipped.
Use --all to export every row regardless of the filter.

BOM handling
------------
Billbee's XLSX format uses numbered column sets per BOM item:
  Subarticle 1 Id / Subarticle 1 SKU / Subarticle 1 Amount
  Subarticle 2 Id / ...
The script builds a SKU→Id map from ALL sheet rows (not just visible ones)
so that BOM references can be resolved even when the physical product rows
are hidden by the current filter.

Usage
-----
  python execution/export_to_billbee_xlsx.py --sheet-url URL
  python execution/export_to_billbee_xlsx.py --sheet-url URL --all
  python execution/export_to_billbee_xlsx.py --sheet-url URL --output my_import.xlsx
  python execution/export_to_billbee_xlsx.py --sheet-url URL --dry-run

Then in Billbee: Artikel → Importieren → Billbee XLSX → select the file.

Column names verified against a real Billbee XLSX export (Mar 2026).
Since the upload tab uses the same column names as the Billbee XLSX (e.g.
"Weight (g) net", "Custom Field Produktkategorie"), no translation is needed.
"""

import argparse
import sys
from datetime import date
from pathlib import Path

import gspread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))

from google_sheets_client import open_sheet

TAB_NAME              = "ProductList"
COLUMNS_TO_UPLOAD_TAB = "ColumnsToUpload"

# Columns that exist in the upload tab but must never be exported to Billbee.
# These are either pipeline-internal or not importable via Billbee XLSX.
_SKIP_COLS: set[str] = {
    "Type", "IsDeactivated", "IsDigital", "IsBom", "Action",
    "BOM_Count", "BOM_SKUs",     # BOM_SKUs → Subarticle columns (generated at export)
    "Sources",                   # collapsed summary; Source 1 * cols handle individual sources
    "Tags",                      # old aggregate; Tags DE is the Billbee column
}

# Always force-include even if not in ColumnsToUpload (Billbee needs these to
# match the correct existing article on import).
_ALWAYS_INCLUDE: set[str] = {"Id", "SKU"}

# Width hints for XLSX column auto-sizing (xlsx header → pixel width)
_WIDTH_HINTS: dict[str, int] = {
    "Id": 18, "SKU": 38, "EAN": 16, "IsBom": 10,
    "Manufacturer": 16,
    "Title DE": 42, "Short description DE": 36, "Long description DE": 60,
    "Materials DE": 30, "Tags DE": 30,
    "Invoice text DE": 40, "Export description DE": 40, "Basic attributes DE": 30,
    "Price gross": 12, "CostPrice gross": 14, "Price net": 12, "CostPrice net": 14,
    "VAT index": 10,
    "Weight (g) net": 13, "Weight (g) gross": 14,
    "LengthCm": 10, "WidthCm": 10, "HeightCm": 10,
    "TARIC Code": 14, "Country of origin": 16,
    "Stock current Standard": 20, "Stock min Standard": 18,
    "Stock target Standard": 20, "Stock place Standard": 18,
    "Custom Field Produktkategorie": 22, "Custom Field Produktgröße": 16,
    "Custom Field Produktvariante": 18, "Custom Field Produktfarbe": 16,
    "Condition": 12, "Units per item": 14, "Unit": 10,
    "Shipping product": 16, "Delivery time": 14,
    "Source 1 Partner": 16, "Source 1 Source Id": 20,
    "Resolve Parts": 13, "Set Split Price": 14, "Status": 10,
}


# ──────────────────────────────────────────────────────────────────────────────
# ColumnsToUpload reader
# ──────────────────────────────────────────────────────────────────────────────

def read_columns_to_upload(spreadsheet) -> list[str]:
    """
    Read the ColumnsToUpload tab and return column names whose checkbox is TRUE,
    in the order they appear in the tab.  Returns an empty list if tab not found.
    """
    try:
        ws   = spreadsheet.worksheet(COLUMNS_TO_UPLOAD_TAB)
        data = ws.get_all_values()
        checked: list[str] = []
        for row in data[1:]:   # skip header
            if row and len(row) >= 2:
                col_name   = row[0].strip()
                is_checked = row[1].strip().upper() == "TRUE"
                if col_name and is_checked:
                    checked.append(col_name)
        return checked
    except gspread.exceptions.WorksheetNotFound:
        return []


# ──────────────────────────────────────────────────────────────────────────────
# Row visibility (Google Sheets filter)
# ──────────────────────────────────────────────────────────────────────────────

def get_visible_row_indices(spreadsheet, tab_name: str) -> set[int]:
    """
    Return 0-based data-row indices that are NOT hidden by a Google Sheets filter
    (index 0 = first data row after the header).
    """
    ws = spreadsheet.worksheet(tab_name)
    result = spreadsheet.client.spreadsheets_get(
        spreadsheet.id,
        params={
            "ranges": f"'{tab_name}'",
            "fields": "sheets(properties(sheetId),data(rowMetadata(hiddenByFilter)))",
            "includeGridData": "false",
        },
    )
    sheet_id   = ws.id
    sheet_data = next(
        (s for s in result.get("sheets", [])
         if s.get("properties", {}).get("sheetId") == sheet_id),
        None,
    )
    if sheet_data is None:
        raise RuntimeError(f"Tab '{tab_name}' not found in Sheets API response")

    row_metadata = sheet_data.get("data", [{}])[0].get("rowMetadata", [])
    return {
        i - 1                                    # 0-based data index
        for i, meta in enumerate(row_metadata)
        if i > 0 and not meta.get("hiddenByFilter", False)
    }


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _int_id(val) -> int | None:
    try:
        return int(float(str(val))) if str(val).strip() else None
    except (ValueError, TypeError):
        return None


def _is_delete(row: dict) -> bool:
    return str(row.get("Action") or "").strip().lower() == "delete"


def _bom_skus(row: dict) -> list[str]:
    raw = str(row.get("BOM_SKUs") or "").strip()
    return [s.strip() for s in raw.split("|") if s.strip()] if raw else []


def _cell_value(row: dict, col: str):
    """Convert a sheet cell value to the XLSX cell value (numeric where possible).

    Exception: 'Source N Stocksync active' columns are kept as strings ('0'/'1')
    to match Billbee's own XLSX export format exactly.
    """
    raw = row.get(col)
    if raw is None or str(raw).strip() in ("", "None"):
        return None
    s = str(raw).strip()
    # Keep Stocksync active as string — Billbee exports '0'/'1', not integers.
    if "Stocksync active" in col:
        return s
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        pass
    return s


# ──────────────────────────────────────────────────────────────────────────────
# XLSX generation
# ──────────────────────────────────────────────────────────────────────────────

def build_xlsx(
    rows: list[dict],
    all_rows: list[dict],
    export_cols: list[str],
) -> openpyxl.Workbook:
    """
    Build a Billbee-import-compatible XLSX workbook.

    rows        — the rows to export (visible / filtered, non-deleted)
    all_rows    — ALL sheet rows, used to build the SKU→Id map for BOM resolution
    export_cols — ordered list of column names to include (from ColumnsToUpload),
                  already filtered to remove skip-cols; always includes Id + SKU
    """
    # ── Build SKU→BillbeeId map from all rows (not just visible) ─────────────
    sku_to_id: dict[str, int] = {}
    for r in all_rows:
        pid = _int_id(r.get("Id"))
        sku = str(r.get("SKU") or "").strip()
        if pid and sku:
            sku_to_id[sku] = pid

    # ── Determine max BOM depth across all export rows ────────────────────────
    max_bom = max((len(_bom_skus(r)) for r in rows), default=0)

    # ── Build final column list ───────────────────────────────────────────────
    # Regular columns: name = both the sheet key AND the XLSX header (they match now)
    # Synthetic BOM columns: prefixed __bom_*
    columns: list[str] = list(export_cols)  # copy; names are direct XLSX headers too

    bom_col_start = len(columns)
    for n in range(1, max_bom + 1):
        columns.append(f"__bom_id_{n}")
        columns.append(f"__bom_sku_{n}")
        columns.append(f"__bom_amt_{n}")

    # ── Create workbook ───────────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Artikel"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center")

    def _xlsx_header(col: str) -> str:
        if col.startswith("__bom_"):
            parts = col.split("_")   # ['', '', 'bom', field, N]
            field = parts[3]
            n     = parts[4]
            if field == "id":  return f"Subarticle {n} Id"
            if field == "sku": return f"Subarticle {n} SKU"
            if field == "amt": return f"Subarticle {n} Amount"
        return col   # sheet column name IS the Billbee XLSX header

    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=_xlsx_header(col))
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"

    # ── Write data rows ───────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 2):
        bom_entries: list[tuple[int | None, str]] = []
        for bom_sku in _bom_skus(row):
            bid = sku_to_id.get(bom_sku)
            if bid is None:
                print(f"  [warn] BOM SKU {bom_sku!r} not found in sheet — Id will be empty")
            bom_entries.append((bid, bom_sku))

        for col_idx, col in enumerate(columns, 1):
            val = None

            if col.startswith("__bom_"):
                parts = col.split("_")
                field = parts[3]
                n     = int(parts[4]) - 1   # 0-based
                if n < len(bom_entries):
                    bid, bsku = bom_entries[n]
                    if field == "id"  and bid is not None: val = bid
                    if field == "sku":                      val = bsku
                    if field == "amt":                      val = 1
            else:
                val = _cell_value(row, col)

            if val is not None:
                ws.cell(row=row_idx, column=col_idx, value=val)

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col in enumerate(columns, 1):
        header = _xlsx_header(col)
        if header.startswith("Subarticle"):
            width = 38 if "SKU" in header else (18 if "Id" in header else 9)
        else:
            width = _WIDTH_HINTS.get(header, 16)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return wb


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export visible Google Sheet rows to a Billbee XLSX import file.",
    )
    parser.add_argument("--sheet-url", required=True,
                        help="URL of the Google Sheet with the 'upload' tab.")
    parser.add_argument("--output",
                        help="Output XLSX path.  Default: billbee_import_YYYY-MM-DD.xlsx")
    parser.add_argument("--all", dest="all_rows", action="store_true",
                        help="Export all rows (ignore Google Sheets filter).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print row count and first few SKUs without writing a file.")
    args = parser.parse_args()

    print("Opening sheet …")
    spreadsheet = open_sheet(args.sheet_url)
    ws          = spreadsheet.worksheet(TAB_NAME)
    all_records = ws.get_all_records()
    print(f"  {len(all_records)} total rows in '{TAB_NAME}' tab.")

    # ── Read ColumnsToUpload ──────────────────────────────────────────────────
    print(f"Reading '{COLUMNS_TO_UPLOAD_TAB}' tab …")
    checked_cols = read_columns_to_upload(spreadsheet)
    if not checked_cols:
        print(f"  [warn] '{COLUMNS_TO_UPLOAD_TAB}' tab not found or no columns checked.")

    # Build export column list: always_include + checked (minus skips) in order
    seen: set[str] = set()
    export_cols: list[str] = []

    # Force Id and SKU first
    for col in ("Id", "SKU"):
        export_cols.append(col)
        seen.add(col)

    for col in checked_cols:
        if col not in seen and col not in _SKIP_COLS:
            export_cols.append(col)
            seen.add(col)

    print(f"  Export columns ({len(export_cols)}): {export_cols}")

    # ── Select rows ───────────────────────────────────────────────────────────
    if args.all_rows:
        export_rows = all_records
        print(f"  --all: exporting all {len(export_rows)} rows.")
    else:
        print("  Reading row visibility (Google Sheets filter) …")
        visible_idx = get_visible_row_indices(spreadsheet, TAB_NAME)
        export_rows = [r for i, r in enumerate(all_records) if i in visible_idx]
        print(f"  {len(export_rows)} visible rows "
              f"({len(all_records) - len(export_rows)} hidden).")

    # Skip Action=delete rows
    before      = len(export_rows)
    export_rows = [r for r in export_rows if not _is_delete(r)]
    if len(export_rows) < before:
        print(f"  Skipped {before - len(export_rows)} Action=delete rows.")

    if not export_rows:
        print("[warn] No rows to export.  Check the filter or use --all.")
        sys.exit(0)

    # BOM stats
    bom_rows = [r for r in export_rows if _bom_skus(r)]
    max_bom  = max((len(_bom_skus(r)) for r in export_rows), default=0)
    print(f"  {len(bom_rows)} listing rows with BOM; max BOM depth = {max_bom}")

    if args.dry_run:
        print(f"\n[DRY-RUN] Would export {len(export_rows)} rows.  First 10 SKUs:")
        for r in export_rows[:10]:
            bom = _bom_skus(r)
            bom_str = f"  BOM: {' | '.join(bom)}" if bom else ""
            print(f"  {r.get('SKU', '(no SKU)')}{bom_str}")
        if len(export_rows) > 10:
            print(f"  … and {len(export_rows) - 10} more")
        return

    print(f"\nBuilding XLSX for {len(export_rows)} rows …")
    wb = build_xlsx(export_rows, all_records, export_cols)

    output_path = Path(args.output) if args.output else \
        Path(f"billbee_import_{date.today().strftime('%Y-%m-%d')}.xlsx")

    wb.save(output_path)
    print(f"\n[done] Saved: {output_path}")
    print()
    print("Next step: Billbee → Artikel → Importieren → Billbee XLSX → select file")


if __name__ == "__main__":
    main()
