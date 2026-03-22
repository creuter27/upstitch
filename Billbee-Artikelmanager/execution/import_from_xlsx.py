"""
Import a Billbee XLSX export into a Google Sheet.

Process
-------
1. Read the Billbee XLSX file (exact column structure, unmodified).
2. Write to 'downloaded' tab — the immutable original. Never touched by pipeline
   scripts.
3. Create 'upload' tab — copy of 'downloaded' plus management columns that the
   pipeline needs:
     Type      derived from IsBom (2 = listing/BOM, 1 = physical)
     BOM_Count count of non-empty Subarticle N SKU values
     BOM_SKUs  pipe-separated list of Subarticle SKUs (pipeline curates these)
     Action    empty — dedup/pipeline scripts set this to 'delete'
4. Create/update 'ColumnsToUpload' tab with checkbox defaults.

Usage
-----
  # Create a new sheet (named "Billbee Artikelmanager YYYY-MM-DD"):
  python execution/import_from_xlsx.py --xlsx-file billbee_export.xlsx

  # Reuse an existing sheet (overwrites downloaded + upload + ColumnsToUpload):
  python execution/import_from_xlsx.py --xlsx-file billbee_export.xlsx --sheet-url URL
"""

import argparse
import re
import sys
from datetime import date
from pathlib import Path

import gspread
import openpyxl
from gspread.utils import rowcol_to_a1

sys.path.insert(0, str(Path(__file__).parent.parent))
from google_sheets_client import create_sheet, open_sheet

COLUMNS_TO_UPLOAD_TAB = "ColumnsToUpload"

# Management columns appended to the 'upload' tab (not present in Billbee XLSX).
# Order matters — this is the order they appear after the Billbee columns.
MANAGEMENT_COLUMNS: list[str] = ["Type", "BOM_Count", "BOM_SKUs", "Action"]

# Columns that should never appear in ColumnsToUpload (internal pipeline state).
_NEVER_UPLOAD: set[str] = {"Type", "BOM_Count", "BOM_SKUs", "Action"}

# Columns pre-checked by default in ColumnsToUpload — the fields our pipeline
# actively modifies.
_DEFAULT_CHECKED: set[str] = {
    "Custom Field Produktkategorie",
    "Custom Field Produktgröße",
    "Custom Field Produktvariante",
    "Custom Field Produktfarbe",
    "TARIC Code",
    "Country of origin",
}


# ── XLSX reading ─────────────────────────────────────────────────────────────

def read_xlsx(path: Path) -> tuple[list[str], list[list[str]]]:
    """
    Read a Billbee XLSX file.
    Returns (headers, data_rows) with all values converted to strings.
    Entirely empty rows are skipped.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not raw_rows:
        raise ValueError("XLSX file is empty")

    headers = [str(h).strip() if h is not None else "" for h in raw_rows[0]]
    n = len(headers)

    data: list[list[str]] = []
    for raw in raw_rows[1:]:
        if not any(v is not None and str(v).strip() for v in raw):
            continue   # skip blank rows
        row = [str(v).strip() if v is not None else "" for v in raw]
        # Pad short rows; truncate over-wide rows
        row = (row + [""] * n)[:n]
        data.append(row)

    return headers, data


# ── Upload-tab construction ──────────────────────────────────────────────────

def _subarticle_skus(headers: list[str], row: list[str]) -> list[str]:
    """
    Collect all non-empty 'Subarticle N SKU' values, sorted by N.
    """
    pairs: list[tuple[int, str]] = []
    for col, val in zip(headers, row):
        m = re.fullmatch(r"Subarticle (\d+) SKU", col)
        if m and val.strip():
            pairs.append((int(m.group(1)), val.strip()))
    pairs.sort()
    return [v for _, v in pairs]


def build_upload_rows(
    headers: list[str],
    data: list[list[str]],
) -> tuple[list[str], list[list[str]]]:
    """
    Extend headers and rows with management columns for the 'upload' tab.

    Added columns:
      Type      "2" if IsBom == TRUE, else "1"
      BOM_Count number of Subarticle SKUs
      BOM_SKUs  pipe-separated Subarticle SKUs
      Action    "" (empty — pipeline fills this)
    """
    upload_headers = headers + MANAGEMENT_COLUMNS
    is_bom_idx = headers.index("IsBom") if "IsBom" in headers else None

    upload_data: list[list[str]] = []
    for row in data:
        is_bom = ""
        if is_bom_idx is not None and is_bom_idx < len(row):
            is_bom = row[is_bom_idx].strip().upper()

        type_val  = "2" if is_bom == "TRUE" else "1"
        sub_skus  = _subarticle_skus(headers, row)
        bom_count = str(len(sub_skus))
        bom_skus  = " | ".join(sub_skus)

        upload_data.append(row + [type_val, bom_count, bom_skus, ""])

    return upload_headers, upload_data


# ── Google Sheets writing ─────────────────────────────────────────────────────

def _write_tab_chunked(
    ws: gspread.Worksheet,
    headers: list[str],
    data: list[list[str]],
    chunk_size: int = 500,
) -> None:
    """Write header row then data in chunks, overwriting existing content."""
    n_cols = len(headers)
    last_col = rowcol_to_a1(1, n_cols)[:-1]   # e.g. "BU" from "BU1"

    ws.update(values=[headers], range_name="A1", value_input_option="RAW")

    total = (len(data) + chunk_size - 1) // chunk_size
    for b, i in enumerate(range(0, len(data), chunk_size), start=1):
        chunk = data[i:i + chunk_size]
        start = i + 2
        end   = start + len(chunk) - 1
        ws.update(
            values=chunk,
            range_name=f"A{start}:{last_col}{end}",
            value_input_option="RAW",
        )
        print(f"    chunk {b}/{total}: rows {start}–{end}")


def write_tab(
    ss: gspread.Spreadsheet,
    tab_name: str,
    headers: list[str],
    data: list[list[str]],
) -> None:
    """Create or clear a worksheet and write headers + data."""
    n_rows = len(data) + 1
    n_cols = len(headers)
    try:
        ws = ss.worksheet(tab_name)
        ws.clear()
        if ws.col_count < n_cols:
            ws.resize(cols=n_cols)
    except gspread.exceptions.WorksheetNotFound:
        ws = ss.add_worksheet(title=tab_name, rows=n_rows + 10, cols=n_cols)

    _write_tab_chunked(ws, headers, data)
    print(f"  [ok] '{tab_name}': {len(data)} rows × {n_cols} columns.")


# ── ColumnsToUpload tab ───────────────────────────────────────────────────────

def update_columns_to_upload_tab(
    ss: gspread.Spreadsheet,
    all_columns: list[str],
) -> None:
    """Create or update the ColumnsToUpload tab."""
    upload_cols = [c for c in all_columns if c not in _NEVER_UPLOAD]

    # Preserve existing checkbox values where possible
    existing: dict[str, bool] = {}
    ws_exists = False
    try:
        ws = ss.worksheet(COLUMNS_TO_UPLOAD_TAB)
        for row in ws.get_all_values()[1:]:
            if row:
                checked = (row[1].strip().upper() == "TRUE") if len(row) > 1 else False
                existing[row[0]] = checked
        ws_exists = True
    except gspread.exceptions.WorksheetNotFound:
        pass

    merged: list[tuple[str, bool]] = [
        (col, existing.get(col, col in _DEFAULT_CHECKED))
        for col in upload_cols
    ]
    n = len(merged)
    matrix = [["Column", "Upload to Billbee"]] + [[name, val] for name, val in merged]

    if ws_exists:
        ws.clear()
    else:
        ws = ss.add_worksheet(title=COLUMNS_TO_UPLOAD_TAB, rows=n + 5, cols=3)

    ws.update(values=matrix, range_name="A1", value_input_option="RAW")

    ss.batch_update({"requests": [
        # Bold header
        {"repeatCell": {
            "range": {"sheetId": ws.id, "startRowIndex": 0, "endRowIndex": 1,
                      "startColumnIndex": 0, "endColumnIndex": 2},
            "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
            "fields": "userEnteredFormat(textFormat)",
        }},
        # Checkbox validation on column B
        {"setDataValidation": {
            "range": {"sheetId": ws.id, "startRowIndex": 1, "endRowIndex": 1 + n,
                      "startColumnIndex": 1, "endColumnIndex": 2},
            "rule": {"condition": {"type": "BOOLEAN"}, "strict": True, "showCustomUi": True},
        }},
        # Column A width
        {"updateDimensionProperties": {
            "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1},
            "properties": {"pixelSize": 260}, "fields": "pixelSize",
        }},
        # Column B width
        {"updateDimensionProperties": {
            "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 2},
            "properties": {"pixelSize": 150}, "fields": "pixelSize",
        }},
    ]})
    checked = sum(1 for _, v in merged if v)
    print(f"  [ok] '{COLUMNS_TO_UPLOAD_TAB}': {n} columns ({checked} pre-checked).")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Import a Billbee XLSX export into a Google Sheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--xlsx-file", required=True, type=Path,
        help="Path to the Billbee XLSX export file.",
    )
    parser.add_argument(
        "--sheet-url",
        help="Reuse an existing Google Sheet (overwrites downloaded/upload/ColumnsToUpload). "
             "If omitted, a new sheet is created.",
    )
    args = parser.parse_args()

    if not args.xlsx_file.exists():
        print(f"[error] File not found: {args.xlsx_file}")
        sys.exit(1)

    # 1 ── Read XLSX
    print(f"[1/4] Reading '{args.xlsx_file.name}' …")
    headers, data = read_xlsx(args.xlsx_file)
    print(f"      {len(headers)} columns, {len(data)} rows.")

    # 2 ── Open or create sheet
    if args.sheet_url:
        print("[2/4] Opening existing sheet …")
        ss = open_sheet(args.sheet_url)
    else:
        title = f"Billbee Artikelmanager {date.today().strftime('%Y-%m-%d')}"
        print(f"[2/4] Creating sheet: '{title}' …")
        ss = create_sheet(title)
    print(f"      '{ss.title}'")

    # 3 ── Write 'downloaded' tab (exact XLSX content, immutable reference)
    print("[3/4] Writing 'downloaded' tab …")
    write_tab(ss, "downloaded", headers, data)

    # 3 ── Write 'ProductList' tab (with management columns; pipeline works here)
    print("[3/4] Writing 'ProductList' tab …")
    upload_headers, upload_data = build_upload_rows(headers, data)
    write_tab(ss, "ProductList", upload_headers, upload_data)

    # 4 ── ColumnsToUpload tab
    print("[4/4] Updating 'ColumnsToUpload' tab …")
    update_columns_to_upload_tab(ss, upload_headers)

    print(f"\n[done]")
    print(f"  Sheet:      {ss.url}")
    print(f"  downloaded: {len(data)} rows  (original XLSX — never modify)")
    print(f"  upload:     {len(upload_data)} rows  (pipeline works here)")
    print()
    print("Next steps:")
    print("  1. Review + check columns in 'ColumnsToUpload' tab")
    print("  2. Run pipeline scripts against --sheet-url (they use the 'upload' tab)")
    print("  3. Export: python execution/export_to_billbee_xlsx.py --sheet-url URL")
    print("  4. Upload the XLSX in Billbee: Artikel → Importieren → Billbee XLSX")


if __name__ == "__main__":
    main()
